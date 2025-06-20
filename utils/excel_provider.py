# utils/excel_provider.py
import os
from pathlib import Path
from openpyxl import load_workbook
from functools import lru_cache

# utils/excel_provider.py (The Engine)
# Purpose: This is the core, reusable library. It does all the heavy lifting: finding the file, reading the data, caching it for performance, and performing the filtering.
# Analogy: Think of this as the powerful engine of a car.

class ExcelDataProvider:
    """
    A reusable utility for reading and filtering data from Excel files.
    Features robust path handling and caching for performance.
    """

    def __init__(self, workbook_name, data_folder="data"):
        """
        Initializes the provider with the workbook name.

        Args:
            workbook_name (str): The name of the Excel file (e.g., "TestData_AppName.xlsx").
            data_folder (str): The name of the folder where test data is stored.
        """
        # Build a reliable, absolute path to the Excel file
        project_root = Path(__file__).parent.parent
        self.file_path = project_root / data_folder / workbook_name

        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel workbook not found at the expected path: {self.file_path}")

    @lru_cache(maxsize=None)  # This is a simple and powerful caching decorator
    def get_sheet_data(self, sheet_name):
        """
        Reads all data from a specific sheet and caches the result.
        The first row is treated as headers.

        Args:
            sheet_name (str): The name of the sheet to read.

        Returns:
            list[dict]: A list of dictionaries representing the rows.
        """
        try:
            workbook = load_workbook(self.file_path, data_only=True)
            worksheet = workbook[sheet_name]
        except KeyError:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{self.file_path.name}'")

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() for cell in rows[0]]
        data = []

        for row in rows[1:]:
            if any(cell is not None for cell in row):  # Skip completely empty rows
                row_dict = {header: value for header, value in zip(headers, row)}
                data.append(row_dict)

        workbook.close()
        return data

    def get_all_data(self, sheet_name):
        """
        Public method to get all data from a sheet. Uses the cached method.

        Args:
            sheet_name (str): The name of the sheet.
        """
        return self.get_sheet_data(sheet_name)

    # Add these methods to the ExcelDataProvider class in utils/excel_provider.py

    def get_data_by_key_value(self, sheet_name, filter_key, filter_value):
        """
        Filters data from a sheet where a specific column has a specific value.

        Args:
            sheet_name (str): The name of the sheet.
            filter_key (str): The column header to filter on.
            filter_value: The value to match.

        Returns:
            list[dict]: A list of matching rows.
        """
        all_data = self.get_sheet_data(sheet_name)

        # Convert filter_value to string for consistent comparison, as Excel data is read as mixed types
        filter_value = str(filter_value)

        return [
            row for row in all_data
            if str(row.get(filter_key)).strip() == filter_value
        ]

    def get_data_with_filters(self, sheet_name, filters):
        """
        Filters data based on multiple AND conditions.

        Args:
            sheet_name (str): The name of the sheet.
            filters (dict): A dictionary of {column: value} pairs to filter by.

        Returns:
            list[dict]: A list of matching rows.
        """
        all_data = self.get_sheet_data(sheet_name)

        if not filters:
            return all_data

        filtered_data = []
        for row in all_data:
            match = True
            for key, value in filters.items():
                if str(row.get(key)).strip() != str(value).strip():
                    match = False
                    break  # Mismatch found, no need to check other filters for this row
            if match:
                filtered_data.append(row)

        return filtered_data